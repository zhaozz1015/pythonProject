#encoding: utf-8
import pandas as pd
import sys
reload(sys)
sys.setdefaultencoding('utf8')
import os
import datetime
from openpyxl import load_workbook

result_file_path = 'rest\\result_' + datetime.datetime.now().__format__('%Y-%m-%d') + '.xlsx'

def log_report(message):
    create_file()

    sheet_name = 'reports'
    df = pd.DataFrame([('', datetime.datetime.now(), message)], columns=['', '时间', '信息'])  # 列表数据转为数据框
    df1 = pd.DataFrame(pd.read_excel(result_file_path, sheet_name=sheet_name))  # 读取原数据文件和表
    writer = pd.ExcelWriter(result_file_path, engine='openpyxl')
    book = load_workbook(result_file_path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df_rows = df1.shape[0]  # 获取原数据的行数
    df.to_excel(writer, sheet_name=sheet_name, startrow=df_rows + 1, index=False, header=False)  # 将数据写入excel中的aa表,从第一个空行开始写
    writer.save()

def log_result(type, message):
    create_file()

    sheet_name = 'problems'
    df = pd.DataFrame([('', datetime.datetime.now(), type, message)], columns=['', '时间', '类型', '消息'])  # 列表数据转为数据框
    df1 = pd.DataFrame(pd.read_excel(result_file_path, sheet_name=sheet_name))  # 读取原数据文件和表
    writer = pd.ExcelWriter(result_file_path, engine='openpyxl')
    book = load_workbook(result_file_path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df_rows = df1.shape[0]  # 获取原数据的行数
    df.to_excel(writer, sheet_name=sheet_name, startrow=df_rows + 1, index=False, header=False)  # 将数据写入excel中的aa表,从第一个空行开始写
    writer.save()

def create_file():
    if (not os.path.exists(result_file_path)):
        writer = pd.ExcelWriter(result_file_path)
        data1 = pd.DataFrame(columns=['时间', '类型', '消息'])
        data1.to_excel(writer, sheet_name='problems')
        data2 = pd.DataFrame(columns=['时间', '信息'])
        data2.to_excel(writer, sheet_name='reports')
        writer.save()